import os

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from python_code.cruds import dish_crud as DC
from python_code.cruds import menu_crud as MC
from python_code.cruds import submenu_crud as SC
from python_code.dao.redis_dao import RedisDAO
from python_code.db import Session
from python_code.logger import main_logger
from python_code.schemas.dish_schemas import BaseDish, DishSchema
from python_code.schemas.dish_schemas import CreateDishWithId
from python_code.schemas.menu_schemas import CreateMenuWithId
from python_code.schemas.submenu_schemas import CreateSubmenuWithId


def read_excel():
    """
    parse excel file and get:
    ~=list[Menu]
    ~=list[Submenu]
    ~=list[Dish]
    """

    # Получаем текущую директорию скрипта
    current_directory = os.getcwd()

    # Строим путь к файлу относительно текущей директории
    excel_filepath = os.path.join(current_directory, 'Menu.xlsx')
    wb = openpyxl.load_workbook(excel_filepath)
    sheet = wb.active

    menus = []
    submenus = []
    dishes = []

    row = 1
    while row <= sheet.max_row:
        if (sheet.cell(row=row, column=1).value
                and type(sheet.cell(row=row, column=1).value) == str):  # Новое меню
            menu = {
                'id': sheet.cell(row=row, column=1).value,
                'title': sheet.cell(row=row, column=2).value,
                'description': sheet.cell(row=row, column=3).value
            }
            menus.append(menu)
            row += 1

            while (row <= sheet.max_row
                   and not type(sheet.cell(row=row, column=1).value) == str):  # Подменю
                if (sheet.cell(row=row, column=2).value
                        and type(sheet.cell(row=row, column=2).value) == str):
                    submenu = {
                        'id': sheet.cell(row=row, column=2).value,
                        'title': sheet.cell(row=row, column=3).value,
                        'description': sheet.cell(row=row, column=4).value,
                        'menu_id': menu['id']
                    }
                    submenus.append(submenu)
                    row += 1

                    while (row <= sheet.max_row
                           and not isinstance(sheet.cell(row=row, column=1).value, str)
                           and not isinstance(sheet.cell(row=row, column=2).value, str)):  # Блюдо
                        base_price_value = sheet.cell(row=row, column=6).value
                        discount = sheet.cell(row=row, column=7).value if type(
                            sheet.cell(row=row, column=7).value) == str else 0.0
                        discount = 1 - float(discount)
                        # price_value = float(base_price_value.replace(',', '.'))
                        price_value = float(base_price_value)
                        price_value = price_value * discount

                        dish = {
                            'id': sheet.cell(row=row, column=3).value,
                            'title': sheet.cell(row=row, column=4).value,
                            'description': sheet.cell(row=row, column=5).value,
                            'price': price_value,
                            'submenu_id': submenu['id'],
                            'menu_id': menu['id']  # Добавляем информацию о меню
                        }
                        dishes.append(dish)
                        row += 1
                else:
                    row += 1
        else:
            row += 1

    return menus, submenus, dishes


async def update_db_from_excel(menus, submenus, dishes):
    async with Session() as session:
        result = await compare_menu(session, menus)
        result.append(await compare_submenu(session, submenus))
        result.append(await compare_dish(session, dishes))
        if result == [[],[]]:
            return "DB already updated"
        return result


async def compare_menu(session: AsyncSession, menus: list[dict]):
    result = []
    for menu in menus:
        valid_data: CreateMenuWithId = CreateMenuWithId.model_validate(menu, strict=False)
        menu_from_db = await MC.get_menu_by_id(menu.get('id'), session)
        if menu_from_db is None:
            r = await MC.create_menu(valid_data, session)
            result.append(r)
        else:
            if (valid_data.title != menu_from_db.title
                    or valid_data.description != menu_from_db.description):
                r = await MC.update_menu_by_id(menu.get('id'), valid_data, session)
                result.append(r)
    return result


async def compare_submenu(session: AsyncSession, submenus: list[dict]):
    result = []
    for submenu in submenus:
        valid_data: CreateSubmenuWithId = CreateSubmenuWithId.model_validate(submenu, strict=False)
        submenu_from_db = await SC.get_submenu_by_id(submenu.get('id'), session)
        if submenu_from_db is None:
            r = await SC.create_submenu(submenu.get('menu_id'), valid_data, session)
            result.append(r)
        else:
            if (valid_data.title != submenu_from_db.title
                    or valid_data.description != submenu_from_db.description):
                r = await SC.update_submenu_by_id(submenu.get('menu_id'), submenu.get('id'), valid_data, session)
                result.append(r)
    return result


async def compare_dish(session: AsyncSession, dishes: list[dict]):
    result = []
    for dish in dishes:
        dish['price'] = round(dish['price'], 2)
        dish['price'] = str(dish['price'])
        valid_data: CreateDishWithId = CreateDishWithId.model_validate(dish, strict=False)
        dish_from_db = await DC.get_dish_by_id(dish.get('id'), session)
        if dish_from_db is None:
            r = await DC.create_dish(dish.get('submenu_id'), valid_data, session)
            result.append(r)
        else:
            if (valid_data.title != dish_from_db.title
                    or valid_data.description != dish_from_db.description
                    or valid_data.price != str(dish_from_db.price)):
                r = await DC.update_dish_by_id(dish.get('submenu_id'), dish.get('id'), valid_data, session)
                result.append(r)
    return result


def round_price(dish: DishSchema | BaseDish | None) -> None:
    'brings the price to  x.xx format'
    if dish:
        dish.price = format(dish.price, '.2f')


async def add_dish_number_to_submenu(submenu, session: AsyncSession) -> None:
    'add dish_count to submenu'
    dishes_count = await SC.count_dishes(submenu.id, session)
    submenu.__setattr__('dishes_count', dishes_count)
    print()


async def add_counters_to_response(menu, session: AsyncSession) -> None:
    'Util func that add counter of dish and sub to menu resp'

    submenus_count = await MC.count_submenu(menu.id, session)
    dishes_count = await MC.count_dishes(menu.id, session)
    menu.__setattr__('submenus_count', submenus_count)
    menu.__setattr__('dishes_count', dishes_count)


async def unvalidate_cache(redis: RedisDAO, path: list[str], request: str = 'unknown'):
    await redis.unvalidate(*path)
    # funcname = inspect.currentframe().f_back.f_code.co_name
    main_logger.info(f'data unvalidated via {request} \nfor listed path: {path}')
